#!/usr/bin/env python3
"""
Apply approved category suggestions from uncategorized_suggestions.xlsx.

Reads the spreadsheet, finds rows with Action="accept", and writes
category overrides to config/category_overrides.json and
taxonomy_review_state.json.

Usage:
    python3 apply_category_suggestions.py [--dry-run]
"""

import json, sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

os_chdir = __import__("os").chdir
os_chdir(Path(__file__).resolve().parent)

XLSX = Path("../uncategorized_suggestions.xlsx")
OVERRIDES_FILE = Path("../config/category_overrides.json")
STATE_FILE = Path("../taxonomy_review_state.json")

# Load taxonomy to get ref mapping
from lxml import etree
tree = etree.parse("../subject-taxonomy-lcsh.xml")
root = tree.getroot()

name_to_ref = {}
ref_to_info = {}
for cat in root.findall("category"):
    for sub in cat.findall("subcategory"):
        for subj in sub.findall("subject"):
            ref = subj.get("ref", "")
            ne = subj.find("name")
            name = ne.text if ne is not None and ne.text else ""
            count = int(subj.get("count", "0"))
            name_to_ref[name] = ref
            ref_to_info[ref] = {
                "name": name, "count": count,
                "from_category": cat.get("label"),
                "from_subcategory": sub.get("label"),
            }

dry_run = "--dry-run" in sys.argv

wb = load_workbook(XLSX, data_only=True)
ws = wb["Category Suggestions"]

accepted = []
excluded = []
skipped = 0
rejected = 0

for row in ws.iter_rows(min_row=2, values_only=False):
    name = row[0].value
    action = str(row[8].value or "").strip().lower()
    approved_cat = str(row[6].value or "").strip()
    approved_sub = str(row[7].value or "").strip()

    if not name:
        continue

    if action == "accept":
        if not approved_cat:
            print(f"  WARNING: '{name}' has action=accept but no approved category, skipping")
            continue
        ref = name_to_ref.get(name)
        if not ref:
            print(f"  WARNING: '{name}' not found in taxonomy, skipping")
            continue
        info = ref_to_info.get(ref, {})
        accepted.append({
            "ref": ref,
            "name": name,
            "count": info.get("count", 0),
            "from_category": info.get("from_category", "Uncategorized"),
            "from_subcategory": info.get("from_subcategory", "General"),
            "to_category": approved_cat,
            "to_subcategory": approved_sub or "General",
        })
    elif action == "exclude":
        ref = name_to_ref.get(name)
        if ref:
            excluded.append({"ref": ref, "name": name})
    elif action == "reject":
        rejected += 1
    else:
        skipped += 1

print(f"Results from spreadsheet:")
print(f"  {len(accepted)} accepted (will create category overrides)")
print(f"  {len(excluded)} excluded (will add to exclusions)")
print(f"  {rejected} rejected (no action)")
print(f"  {skipped} skipped/blank")

if dry_run:
    print("\n[DRY RUN] No files modified.")
    for a in accepted[:10]:
        print(f"  {a['name']} -> {a['to_category']} > {a['to_subcategory']}")
    if len(accepted) > 10:
        print(f"  ... and {len(accepted) - 10} more")
    sys.exit(0)

if not accepted and not excluded:
    print("Nothing to apply.")
    sys.exit(0)

# Load existing overrides
existing_overrides = []
if OVERRIDES_FILE.exists():
    with open(OVERRIDES_FILE) as f:
        existing_overrides = json.load(f)

existing_refs = {o["ref"] for o in existing_overrides}
new_overrides = [a for a in accepted if a["ref"] not in existing_refs]
all_overrides = existing_overrides + new_overrides

with open(OVERRIDES_FILE, "w") as f:
    json.dump(all_overrides, f, indent=2, ensure_ascii=False)
    f.write("\n")
print(f"Wrote {len(all_overrides)} overrides to {OVERRIDES_FILE} ({len(new_overrides)} new)")

# Update taxonomy_review_state.json
state = {}
if STATE_FILE.exists():
    with open(STATE_FILE) as f:
        state = json.load(f)

if "category_overrides" not in state:
    state["category_overrides"] = {}
for a in accepted:
    state["category_overrides"][a["ref"]] = a

if excluded:
    if "exclusions" not in state:
        state["exclusions"] = {}
    for e in excluded:
        state["exclusions"][e["ref"]] = {"name": e["name"]}

state["saved"] = datetime.now().isoformat()
with open(STATE_FILE, "w") as f:
    json.dump(state, f, indent=2, ensure_ascii=False)
print(f"Updated {STATE_FILE}")
