#!/usr/bin/env python3
"""Generate category suggestions for uncategorized taxonomy subjects."""

import json, re
from lxml import etree
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

tree = etree.parse("../subject-taxonomy-lcsh.xml")
root = tree.getroot()

ref_to_cat, ref_to_name = {}, {}
uncategorized, categorized = set(), set()
categories, subcategories = set(), {}

for cat in root.findall("category"):
    cl = cat.get("label")
    categories.add(cl)
    subcategories[cl] = []
    for sub in cat.findall("subcategory"):
        sl = sub.get("label")
        subcategories[cl].append(sl)
        for subj in sub.findall("subject"):
            ref = subj.get("ref", "")
            ne = subj.find("name")
            name = ne.text if ne is not None and ne.text else ""
            ref_to_cat[ref] = (cl, sl)
            ref_to_name[ref] = name
            count = int(subj.get("count", "0"))
            if cl == "Uncategorized":
                uncategorized.add(ref)
            else:
                categorized.add(ref)

# Load appearances
with open("../document_appearances.json") as f:
    appearances = json.load(f)

# Build doc->categorized_refs index
doc_to_refs = {}
for ref in categorized:
    if ref not in appearances:
        continue
    for vol_id, doc_ids in appearances[ref].items():
        for doc_id in doc_ids:
            k = f"{vol_id}/{doc_id}"
            if k not in doc_to_refs:
                doc_to_refs[k] = set()
            doc_to_refs[k].add(ref)

# Keyword-based heuristics
KEYWORD_HINTS = {
    "Arms Control and Disarmament": [
        r"missile", r"warhead", r"ICBM", r"SLBM", r"ABM", r"SALT", r"START",
        r"nuclear", r"weapon", r"bomb", r"launcher", r"verification",
        r"arms\b", r"disarmament", r"treaty", r"nonproliferation", r"fissile",
        r"enrichment", r"centrifuge", r"plutonium", r"uranium", r"megaton",
        r"SS-\d", r"cruise missile", r"ballistic", r"MIRV", r"throw.weight",
        r"reentry vehicle", r"silo", r"telemetry",
    ],
    "Foreign Economic Policy": [
        r"trade", r"tariff", r"export", r"import", r"economic", r"GDP",
        r"inflation", r"currency", r"debt", r"loan", r"credit", r"bank",
        r"commodity", r"oil\b", r"gas\b", r"petroleum", r"price", r"subsid",
        r"investment", r"development", r"agriculture", r"crop", r"wheat",
        r"rice", r"cotton", r"coffee", r"sugar", r"fertilizer", r"mining",
        r"industry", r"manufactur", r"interest rate", r"balance of pay",
        r"deficit", r"surplus", r"sanction", r"embargo", r"exchange rate",
        r"mineral", r"copper", r"iron", r"zinc", r"gold", r"silver",
        r"aluminum", r"tin\b", r"rubber", r"fuel", r"refiner", r"energy",
    ],
    "Politico-Military Issues": [
        r"aircraft", r"helicopter", r"fighter", r"bomber", r"transport",
        r"fleet", r"naval", r"navy", r"army", r"military", r"force",
        r"base\b", r"deploy", r"contingency", r"covert", r"intelligence",
        r"command", r"defense", r"weapon system", r"ammunition", r"patrol",
        r"reconnaissance", r"surveillance", r"radar", r"F-\d", r"C-\d",
        r"A-\d", r"B-\d", r"KC-", r"RH-", r"MIG", r"tank\b",
        r"battalion", r"division", r"regiment", r"brigade",
        r"assassination", r"coup", r"insurgent", r"guerrilla",
    ],
    "Human Rights": [
        r"refugee", r"asylum", r"prisoner", r"torture", r"disappear",
        r"detention", r"human rights", r"freedom", r"persecution",
        r"minority", r"ethnic", r"religious", r"discrimination",
        r"mortality", r"sterilization", r"capital punishment",
        r"death penalty", r"execution",
    ],
    "Global Issues": [
        r"narcotics", r"drug\b", r"cocaine", r"heroin", r"morphine",
        r"opium", r"trafficking", r"immigration", r"population",
        r"environment", r"pollution", r"climate", r"disaster",
        r"epidemic", r"disease", r"health", r"space\b", r"satellite",
        r"election", r"democracy", r"border",
    ],
    "Bilateral Relations": [
        r"bilateral", r"channel", r"summit", r"Dobrynin", r"Gromyko",
        r"Shultz", r"Kissinger", r"ambassador",
    ],
    "Warfare": [
        r"war\b", r"conflict", r"combat", r"invasion", r"occupation",
        r"ceasefire", r"armistice", r"prisoner of war", r"MIA",
        r"casualties", r"hostilities",
    ],
    "International Law": [
        r"law of the sea", r"treaty", r"convention", r"jurisdiction",
        r"sovereignty", r"territorial", r"maritime",
        r"fishing", r"vessel", r"stateless",
    ],
    "Information Programs": [
        r"broadcast", r"radio", r"propaganda", r"exchange",
        r"cultural", r"bookmobile", r"library", r"performing arts",
        r"USIA", r"Voice of America",
    ],
    "Science and Technology": [
        r"nuclear engineering", r"reactor", r"atomic", r"scientific",
        r"technology", r"research", r"CANDU",
    ],
    "Department of State": [
        r"State Department", r"Foreign Service", r"embassy",
        r"consulate", r"diplomatic",
    ],
    "International Organizations": [
        r"United Nations", r"UN\b", r"OAS", r"IMF", r"GATT",
        r"NATO", r"ASEAN",
    ],
}

results = []
for ref in uncategorized:
    name = ref_to_name.get(ref, ref)
    count_xml = 0
    for cat in root.findall("category"):
        if cat.get("label") != "Uncategorized":
            continue
        for sub in cat.findall("subcategory"):
            for subj in sub.findall("subject"):
                if subj.get("ref") == ref:
                    count_xml = int(subj.get("count", "0"))

    # Co-occurrence
    cat_counts = Counter()
    subcat_counts = Counter()
    if ref in appearances:
        for vol_id, doc_ids in appearances[ref].items():
            for doc_id in doc_ids:
                k = f"{vol_id}/{doc_id}"
                for co_ref in doc_to_refs.get(k, set()):
                    c, s = ref_to_cat[co_ref]
                    cat_counts[c] += 1
                    subcat_counts[(c, s)] += 1

    total_co = sum(cat_counts.values())
    if total_co > 0:
        top_cat_co = cat_counts.most_common(1)[0]
        top_sub_co = subcat_counts.most_common(1)[0]
        co_conf = top_cat_co[1] / total_co
        co_cat = top_cat_co[0]
        co_sub = top_sub_co[0][1]
    else:
        co_conf, co_cat, co_sub = 0, None, None

    # Keyword hints
    kw_scores = Counter()
    for cat_label, patterns in KEYWORD_HINTS.items():
        for p in patterns:
            if re.search(p, name, re.IGNORECASE):
                kw_scores[cat_label] += 1
    
    kw_cat = kw_scores.most_common(1)[0][0] if kw_scores else None
    kw_strength = kw_scores.most_common(1)[0][1] if kw_scores else 0

    # Combine signals
    if kw_cat and co_cat and kw_cat == co_cat:
        suggested_cat = co_cat
        suggested_sub = co_sub
        method = "co-occurrence + keyword"
        confidence = min(co_conf + 0.15, 1.0)
    elif kw_strength >= 2:
        suggested_cat = kw_cat
        suggested_sub = subcategories.get(kw_cat, ["General"])[0]
        method = "keyword (strong)"
        confidence = 0.7
    elif co_conf >= 0.5 and total_co >= 10:
        suggested_cat = co_cat
        suggested_sub = co_sub
        method = "co-occurrence"
        confidence = co_conf
    elif kw_strength >= 1 and co_conf >= 0.3:
        suggested_cat = kw_cat
        suggested_sub = co_sub if co_cat == kw_cat else subcategories.get(kw_cat, ["General"])[0]
        method = "keyword + weak co-occurrence"
        confidence = max(co_conf, 0.4)
    elif kw_strength >= 1:
        suggested_cat = kw_cat
        suggested_sub = subcategories.get(kw_cat, ["General"])[0]
        method = "keyword only"
        confidence = 0.35
    elif co_conf >= 0.4 and total_co >= 5:
        suggested_cat = co_cat
        suggested_sub = co_sub
        method = "co-occurrence (weak)"
        confidence = co_conf
    else:
        suggested_cat = None
        suggested_sub = None
        method = "no signal"
        confidence = 0

    # Top 3 co-occurring categories for context
    top3 = cat_counts.most_common(3)
    co_detail = "; ".join(f"{c}: {n}" for c, n in top3) if top3 else ""

    results.append({
        "ref": ref, "name": name, "count": count_xml,
        "suggested_cat": suggested_cat, "suggested_sub": suggested_sub,
        "confidence": confidence, "method": method,
        "co_detail": co_detail, "total_co": total_co,
    })

results.sort(key=lambda x: (-x["confidence"], -x["count"]))

# Build spreadsheet
wb = Workbook()
ws = wb.active
ws.title = "Category Suggestions"

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
HEADER_FILL = PatternFill("solid", fgColor="205493")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(size=11, name="Arial")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D6D7D9"),
)

HIGH_FILL = PatternFill("solid", fgColor="E8F5E9")
MED_FILL = PatternFill("solid", fgColor="FFF8E1")
LOW_FILL = PatternFill("solid", fgColor="FFF3E0")
NONE_FILL = PatternFill("solid", fgColor="FFEBEE")

headers = [
    "Subject Name", "Occurrences", "Suggested Category", "Suggested Subcategory",
    "Confidence", "Method", "Approved Category", "Approved Subcategory",
    "Action", "Co-occurrence Detail",
]
col_widths = [35, 12, 30, 28, 12, 22, 30, 28, 14, 50]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col)].width = w

ws.auto_filter.ref = f"A1:J{len(results)+1}"
ws.freeze_panes = "A2"

# Category dropdown list on a hidden sheet
cat_ws = wb.create_sheet("_Categories")
sorted_cats = sorted(categories - {"Uncategorized"})
for i, c in enumerate(sorted_cats, 1):
    cat_ws.cell(row=i, column=1, value=c)
cat_ws.sheet_state = "hidden"

from openpyxl.worksheet.datavalidation import DataValidation
cat_dv = DataValidation(type="list", formula1=f"=_Categories!$A$1:$A${len(sorted_cats)}", allow_blank=True)
cat_dv.error = "Pick a valid category"
cat_dv.errorTitle = "Invalid Category"
ws.add_data_validation(cat_dv)

action_dv = DataValidation(type="list", formula1='"accept,reject,skip,exclude"', allow_blank=True)
ws.add_data_validation(action_dv)

for row_idx, r in enumerate(results, 2):
    ws.cell(row=row_idx, column=1, value=r["name"]).font = DATA_FONT
    ws.cell(row=row_idx, column=2, value=r["count"]).font = DATA_FONT
    ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row_idx, column=3, value=r["suggested_cat"] or "").font = DATA_FONT
    ws.cell(row=row_idx, column=4, value=r["suggested_sub"] or "").font = DATA_FONT
    
    conf_cell = ws.cell(row=row_idx, column=5, value=r["confidence"])
    conf_cell.font = DATA_FONT
    conf_cell.number_format = "0%"
    conf_cell.alignment = Alignment(horizontal="center")
    
    ws.cell(row=row_idx, column=6, value=r["method"]).font = DATA_FONT
    
    # Pre-fill approved columns with suggestion for high confidence
    if r["confidence"] >= 0.6 and r["suggested_cat"]:
        ws.cell(row=row_idx, column=7, value=r["suggested_cat"]).font = Font(size=11, name="Arial", color="0000FF")
        ws.cell(row=row_idx, column=8, value=r["suggested_sub"]).font = Font(size=11, name="Arial", color="0000FF")
        ws.cell(row=row_idx, column=9, value="accept").font = Font(size=11, name="Arial", color="008000")
    else:
        ws.cell(row=row_idx, column=7, value="").font = DATA_FONT
        ws.cell(row=row_idx, column=8, value="").font = DATA_FONT
        ws.cell(row=row_idx, column=9, value="").font = DATA_FONT
    
    cat_dv.add(ws.cell(row=row_idx, column=7))
    action_dv.add(ws.cell(row=row_idx, column=9))
    
    ws.cell(row=row_idx, column=10, value=r["co_detail"]).font = Font(size=10, name="Arial", color="666666")
    
    # Row coloring by confidence
    if r["confidence"] >= 0.6:
        fill = HIGH_FILL
    elif r["confidence"] >= 0.4:
        fill = MED_FILL
    elif r["confidence"] > 0:
        fill = LOW_FILL
    else:
        fill = NONE_FILL
    
    for col in range(1, 11):
        ws.cell(row=row_idx, column=col).fill = fill
        ws.cell(row=row_idx, column=col).border = THIN_BORDER

# Summary sheet
summary = wb.create_sheet("Summary", 0)
summary_data = [
    ("Category Suggestion Summary", ""),
    ("", ""),
    ("Total uncategorized subjects", len(results)),
    ("High confidence (≥60%)", sum(1 for r in results if r["confidence"] >= 0.6)),
    ("Medium confidence (40-60%)", sum(1 for r in results if 0.4 <= r["confidence"] < 0.6)),
    ("Low confidence (<40%)", sum(1 for r in results if 0 < r["confidence"] < 0.4)),
    ("No suggestion", sum(1 for r in results if r["confidence"] == 0)),
    ("", ""),
    ("Instructions", ""),
    ("1. Review the 'Category Suggestions' sheet", ""),
    ("2. For each row, check the suggested category", ""),
    ("3. Edit 'Approved Category' and 'Approved Subcategory' columns as needed", ""),
    ("4. Set 'Action' to: accept, reject, skip, or exclude", ""),
    ("5. High-confidence rows (green) are pre-filled — just verify", ""),
    ("6. Save and run the apply script to update the taxonomy", ""),
    ("", ""),
    ("Color Key", ""),
    ("Green rows", "High confidence (≥60%) — pre-filled, verify and approve"),
    ("Yellow rows", "Medium confidence (40-60%) — review carefully"),
    ("Orange rows", "Low confidence (<40%) — manual categorization needed"),
    ("Red rows", "No suggestion — categorize manually or exclude"),
]

for row_idx, (a, b) in enumerate(summary_data, 1):
    ca = summary.cell(row=row_idx, column=1, value=a)
    cb = summary.cell(row=row_idx, column=2, value=b)
    ca.font = Font(size=11, name="Arial")
    cb.font = Font(size=11, name="Arial")

summary.cell(row=1, column=1).font = Font(bold=True, size=14, name="Arial", color="205493")
for r in [9]:
    summary.cell(row=r, column=1).font = Font(bold=True, size=12, name="Arial")
for r in [17]:
    summary.cell(row=r, column=1).font = Font(bold=True, size=12, name="Arial")

summary.column_dimensions["A"].width = 45
summary.column_dimensions["B"].width = 55

# Color key cells
for r, fill in [(18, HIGH_FILL), (19, MED_FILL), (20, LOW_FILL), (21, NONE_FILL)]:
    summary.cell(row=r, column=1).fill = fill
    summary.cell(row=r, column=2).fill = fill

out_path = "../uncategorized_suggestions.xlsx"
wb.save(out_path)
print(f"Wrote {out_path}")
print(f"  {len(results)} subjects")
print(f"  {sum(1 for r in results if r['confidence'] >= 0.6)} high confidence (pre-filled)")
print(f"  {sum(1 for r in results if 0.4 <= r['confidence'] < 0.6)} medium confidence")
print(f"  {sum(1 for r in results if 0 < r['confidence'] < 0.4)} low confidence")
print(f"  {sum(1 for r in results if r['confidence'] == 0)} no suggestion")
